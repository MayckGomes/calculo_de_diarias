from openpyxl import *

lista_nomes = []

def mostrar_saldo(saldo_placas: dict):
    print("==" * 50)
    print("SALDOS")
    print("\n")

    acumulador = 0

    for saldo in saldo_placas.items():
        placasaldo = saldo[0]
        valorsaldo = saldo[1]

        acumulador += valorsaldo

        print(f"placa: {placasaldo}, saldo: {valorsaldo}")

    print("=-" * 50)
    print(f"Total = {acumulador}")



def buscar_colaborador(nome: str):

    arquivo = load_workbook("calculo_diarias.xlsx")

    planilha = arquivo["Sheet1"]

    lista_motoristas = planilha["B"]

    lista_ajudantes1 = planilha["E"]

    lista_ajudantes2 = planilha["F"]



    for motorista in range(3, len(lista_motoristas) + 1):

        nome_motorista = planilha.cell(row=motorista, column=2).value

        if nome_motorista == None:
            continue

        if nome in nome_motorista:
            
            if "AGREGADO" in nome_motorista:
                placa = "AGREGADO"
            else:
                placa = planilha.cell(row=motorista, column=3).value

            lista_nomes.append({nome_motorista: placa})



    for ajudante1 in range(3, len(lista_ajudantes1) + 1):

        nome_ajudante1 = planilha.cell(row=ajudante1, column=5).value

        if nome_ajudante1 is None:
            continue

        if nome in nome_ajudante1:
            
            if "AGREGADO" in planilha.cell(row=ajudante1, column=2).value: 

                placa = "AGREGADO"
            
            else:

                placa = planilha.cell(row=ajudante1, column=3).value


            lista_nomes.append({nome_ajudante1: placa})


            

    for ajudante2 in range(3, len(lista_ajudantes2) + 1):

        nome_ajudante2 = planilha.cell(row=ajudante2, column=6).value

        if nome_ajudante2 is None:
            continue

        if nome in nome_ajudante2:


            if "AGREGADO" in planilha.cell(row=ajudante2, column=2).value: 

                placa = "AGREGADO"
            
            else:

                placa = planilha.cell(row=ajudante2, column=3).value
            

            lista_nomes.append({nome_ajudante2: placa})

    return lista_nomes



def verificar_lista(lista: list):

    if len(lista) == 0:
        
        print("Não foi encontrado um colaborador com esse nome!")

        input()
        return {}

    elif len(lista) == 1:

        nome = list(lista[0].keys())[0]
        placa = list(lista[0].values())[0]

        lista.clear()

        print("\n")
        print(f"nome: {nome}, placa: {placa}")

        return {nome: placa}


    if len(lista) > 1:

        while True:

            try:

                print("=" * 100)

                print("Existe mais de um colaborador com este nome, qual seria o certo: ")
                print("\n")
                
                for id_nome in range(len(lista)):

                    nomebusca = list(lista[id_nome].keys())[0]

                    print(f"id: {id_nome}, nome: {nomebusca}")

                input_id = int(input("id do colaborador: "))

                nome = list(lista[input_id].keys())[0]
                placa = list(lista[input_id].values())[0]

                lista.clear()

                print(f"nome: {nome}, placa: {placa}")

                return {nome: placa}

            except:
                print("=-" * 50)
                print("id invalido")
                print("=-" * 50)



def adicionar_saldo(valor: str, colaborador: dict, saldo_list: list):
    print("==" * 50)

    if "," in valor:
        valor = valor.replace(",", ".")
        valor = float(valor)

    else:
        valor = float(valor)

    placa = list(colaborador.values())[0]

    if placa in saldo_list:

        saldo_list[placa] += valor
    else:
        saldo_list[placa] = valor

    mostrar_saldo(saldo_list)